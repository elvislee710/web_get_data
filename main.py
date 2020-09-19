import os
import openpyxl
from web_data import get_data
import send_email

os.chdir(r'C:\\Users\alanlee\Desktop\Future_Product')
print(os.listdir())
path = '//*[@id="calendarFuturesProductTable1"]/tbody/tr[1]/td[3]'
get_list = []
get_num = []
product = []
trade_day = []
wb = openpyxl.load_workbook('Cme_First_Stage.xlsx')
link = 0
try:
    with open('last_trade_day.txt', 'w+') as f:
        f.write("")
    for i in range(6, wb['product_list'].max_row + 1):
        if wb['product_list'].cell(row=i, column=4).value is not None:
            product.append(wb['product_list'].cell(row=i, column=3).value)
            get_list.append(wb['product_list'].cell(row=i, column=4).value)
            get_num.append(i)

    for url1 in get_list:
        settlement = get_data(url1, path)
        print(product[link] + " : " + settlement)
        with open('last_trade_day.txt', 'a+') as f:
            f.write(product[link] + " : " + settlement + '\n')
        link += 1
except Exception as result:
    print(result)

password = input("please Type your password and press enter:  ")
send_email.send_out_email(password)

# home = "https://www.cmegroup.com/trading/energy/crude-oil/light-sweet-crude_product_calendar_futures.html"
# driver = webdriver.Chrome(executable_path=r"C:\Users\alanlee\AppData\Local\Programs\Python\Python38\chromedriver.exe")
# driver.get(home)
# last_trade = driver.find_element_by_xpath('//*[@id="calendarFuturesProductTable1"]/tbody/tr[1]/td[3]').text
# print(last_trade)
