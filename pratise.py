import os
import openpyxl
from web_data import get_data
import send_email

get_list = []
session = []
product = []
path = []
link = 0
time_msg='Check Directory trading_hour.txt in your computer\n'
password = input("please Type your password and press enter:  ")
os.chdir(r'C:\\Users\alanlee\Desktop\Future_Product')
print(os.listdir())
wb = openpyxl.load_workbook('Cme_First_Stage.xlsx')
link = 0
with open('trading_hour.txt', 'w+') as f:
    f.write("")
for i in range(43, wb['product_list'].max_row + 1):
    if wb['product_list'].cell(row=i, column=7).value is not None:
        product.append(wb['product_list'].cell(row=i, column=3).value)
        get_list.append(
            [wb['product_list'].cell(row=i, column=7).value, wb['product_list'].cell(row=i, column=8).value])
        # path.append(wb['product_list'].cell(row=i, column=8).value)
for url in get_list:
    trading_hour = get_data(url[0], url[1])
    session = trading_hour.split('\n')

    if (len(session) >= 2):
        print(product[link] + " : " + session[0] + session[1])
        with open('trading_hour.txt', 'a+') as f:
            f.write(product[link] + " : " + session[0] + session[1] + '\n')
    else:
        print(product[link] + ":" + session[0])
        with open('trading_hour.txt', 'a+') as f:
            f.write(product[link] + " : " + session[0] + '\n')
    link += 1

password = input("please Type your password and press enter:  ")
send_email.send_out_email(password)

# from urllib.request import urlopen
# from bs4 import BeautifulSoup
#
# import random
#
# links = ['http://guba.eastmoney.com']
#
#
# def get_all_link(url):
#     html = urlopen(url)
#     bso = BeautifulSoup(html, 'html.parser')
#     for link in bso.findAll('a'):
#         if 'href' in link.attrs:
#             links.append(link.attrs['href'])
#
#
# get_all_link('http://guba.eastmoney.com')
# print(links)
#
# while (len(links) > 0):
#     links = [link for link in links if link is not None and 'http' in link]
#     print(links)
#     new_page_url = links[random.randint(0, len(links) - 1)]
#     print(new_page_url)
#     get_all_link(new_page_url)
#     print(links)
#     break
